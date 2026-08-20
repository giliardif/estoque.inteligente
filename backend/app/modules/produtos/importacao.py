"""
Etapa 26: parsing de planilha (XLSX/CSV) para import em massa de produtos.

Funções puras aqui — sem acesso a banco. Só transformam bytes de arquivo em
linhas cruas (dict por linha, 1-indexadas a partir da linha 2, já que a
linha 1 é o cabeçalho). Validação de negócio (SKU duplicado, categoria
existente, etc.) fica em service.py, que tem acesso ao tenant.
"""
import csv
import io
from typing import Any

from fastapi import HTTPException, status
from openpyxl import load_workbook

COLUNAS_ESPERADAS = [
    "nome", "sku", "categoria", "codigo_barras", "unidade_medida",
    "custo_medio", "preco_venda", "marca", "ncm", "estoque_minimo", "estoque_maximo",
]


def _normalizar_cabecalho(valor: Any) -> str:
    return str(valor or "").strip().lower().replace(" ", "_")


def _parse_decimal(valor: Any) -> float | None:
    """
    Aceita número já convertido (célula XLSX numérica) ou texto em formato
    BR (1.234,56) ou internacional (1234.56). None/string vazia -> None.
    """
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto:
        return None
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError as exc:
        raise ValueError(f"Valor numérico inválido: '{texto}'.") from exc


def _linhas_de_dicts(cabecalhos: list[str], linhas_brutas: list[list[Any]]) -> list[dict]:
    indices = {nome: i for i, nome in enumerate(cabecalhos) if nome in COLUNAS_ESPERADAS}
    resultado = []
    for numero_linha, linha in enumerate(linhas_brutas, start=2):  # linha 1 = cabeçalho
        if all(c is None or str(c).strip() == "" for c in linha):
            continue  # ignora linhas totalmente vazias (comuns no fim de planilhas)
        dados = {"linha": numero_linha}
        for coluna in COLUNAS_ESPERADAS:
            idx = indices.get(coluna)
            valor = linha[idx] if idx is not None and idx < len(linha) else None
            dados[coluna] = valor.strip() if isinstance(valor, str) else valor
        resultado.append(dados)
    return resultado


def parsear_xlsx(conteudo: bytes) -> list[dict]:
    try:
        workbook = load_workbook(filename=io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Não foi possível ler o arquivo XLSX. Verifique se ele não está corrompido.",
        ) from exc

    planilha = workbook.worksheets[0]
    linhas_iter = planilha.iter_rows(values_only=True)
    try:
        cabecalho_bruto = next(linhas_iter)
    except StopIteration:
        return []
    cabecalhos = [_normalizar_cabecalho(c) for c in cabecalho_bruto]
    linhas_brutas = [list(linha) for linha in linhas_iter]
    workbook.close()
    return _linhas_de_dicts(cabecalhos, linhas_brutas)


def parsear_csv(conteudo: bytes) -> list[dict]:
    try:
        texto = conteudo.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = conteudo.decode("latin-1")

    primeira_linha = texto.splitlines()[0] if texto.splitlines() else ""
    delimitador = ";" if primeira_linha.count(";") > primeira_linha.count(",") else ","

    leitor = csv.reader(io.StringIO(texto), delimiter=delimitador)
    try:
        cabecalho_bruto = next(leitor)
    except StopIteration:
        return []
    cabecalhos = [_normalizar_cabecalho(c) for c in cabecalho_bruto]
    linhas_brutas = [list(linha) for linha in leitor]
    return _linhas_de_dicts(cabecalhos, linhas_brutas)


def parsear_planilha(*, nome_arquivo: str, conteudo: bytes) -> list[dict]:
    nome_lower = (nome_arquivo or "").lower()
    if nome_lower.endswith(".xlsx"):
        return parsear_xlsx(conteudo)
    if nome_lower.endswith(".csv"):
        return parsear_csv(conteudo)
    raise HTTPException(
        status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
        detail="Formato de arquivo não suportado. Envie um arquivo .xlsx ou .csv.",
    )
